#!/usr/bin/python3

import random
random.seed("Christelle")

nr_dossiers = 142
nr_rapporteurs = 9
nr_rapporteurs_par_dossier = 2

nr_dossiers_par_rapporteur_base = (nr_dossiers*nr_rapporteurs_par_dossier) // nr_rapporteurs
rapporteurs_avec_extra = (nr_dossiers*nr_rapporteurs_par_dossier) % nr_rapporteurs

nr_dossiers_pour_rapporteurs = [((nr_dossiers_par_rapporteur_base+1) if rapporteur < rapporteurs_avec_extra else nr_dossiers_par_rapporteur_base) for rapporteur in range(nr_rapporteurs)]
while True:
    try:
        disponibilites = list(nr_dossiers_pour_rapporteurs)
        choix=[]
        for dossier in range(nr_dossiers):
            rapporteurs_disponibles = [i for i in range(nr_rapporteurs) if disponibilites[i]>0]
            rapporteurs = random.sample(rapporteurs_disponibles, nr_rapporteurs_par_dossier)
            for i in rapporteurs:
                disponibilites[i] -= 1
            choix.append(rapporteurs)
    except ValueError:
        pass
    else:
        break

assignation = [[] for i in range(nr_rapporteurs)]
for i in range(nr_dossiers):
    for j in choix[i]:
        assignation[j].append(i)
        
from openpyxl import Workbook
wb = Workbook(write_only=True)

ws_candidats = wb.create_sheet("candidats")
row=["Nom", "Prénom"]
for j in range(nr_rapporteurs_par_dossier):
    row.append("rapport %d"%(j+1))
ws_candidats.append(row)

for i in range(nr_dossiers):
    row = ["nom%03d"%(i+1), "prénom%03d"%(i+1)]
    for j in range(nr_rapporteurs_par_dossier):
        row.append("=rapporteurs!A%d" % (choix[i][j]+2))
    ws_candidats.append(row)
    
ws_rapporteurs = wb.create_sheet("rapporteurs")
row=["Nom"]
ws_rapporteurs.append(row)
for i in range(nr_rapporteurs):
    row = ["rapporteur%02d"%(i+1)]
    for candidat in assignation[i]:
        row.append("=candidats!A%d" % (candidat+2))
        row.append("=candidats!B%d" % (candidat+2))
    ws_rapporteurs.append(row)

wb.save("rapporteurs.xlsx")
